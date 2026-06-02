#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
除濕機全球資料每日自動收集腳本
資料來源：品牌官網、B2B 採購平台（無需登入、穩定可靠）
支援本機執行與 GitHub Actions 雲端排程。
"""

import os
import re
import time
import logging
import smtplib
from pathlib import Path
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 設定 ──────────────────────────────────────────────────────────────────────
OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", Path.home() / "除濕機資料"))
LOG_FILE   = OUTPUT_DIR / "run_log.txt"

SMTP_USER = os.environ.get("GMAIL_USER", "")
SMTP_PASS = os.environ.get("GMAIL_APP_PASS", "")
RECIPIENT = os.environ.get("RECIPIENT_EMAIL", "tc@mjauto.com.tw")

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

COLUMNS = ['生產地', '品牌', '上市時間', '型號', '品名', '主要規格', '單價', '來源', '連結']

# ── 工具函式 ──────────────────────────────────────────────────────────────────
def setup_logging():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=str(LOG_FILE),
        level=logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
        encoding='utf-8',
    )

def safe_get(url, params=None, timeout=25, extra_headers=None):
    h = {**HEADERS, **(extra_headers or {})}
    try:
        r = requests.get(url, params=params, headers=h, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        logging.warning(f"GET 失敗 {url}: {e}")
        return None

def clean(text: str) -> str:
    return re.sub(r'\s+', ' ', text).strip() if text else ''

# ── 台灣市場 ───────────────────────────────────────────────────────────────────

def fetch_pchome_taiwan():
    """PChome 24h（JSON API，最穩定）"""
    products = []
    for page in range(1, 4):
        r = safe_get(
            "https://ecshweb.pchome.com.tw/search/v3.3/all/results",
            params={'q': '除濕機', 'page': page, 'sort': 'new'},
        )
        if not r:
            break
        try:
            data = r.json()
        except Exception:
            break
        for p in data.get('prods', []):
            price_raw = p.get('price', {})
            price = price_raw.get('e', price_raw.get('m', '')) if isinstance(price_raw, dict) else (price_raw or '')
            products.append({
                '生產地': '台灣/亞洲',
                '品牌':   p.get('BrandName', '') or p.get('brand', ''),
                '上市時間': (p.get('publishDate') or '')[:10],
                '型號':   p.get('Id', ''),
                '品名':   clean(p.get('name', '')),
                '主要規格': '',
                '單價':   f"NT$ {price}" if price else '',
                '來源':   'PChome 台灣',
                '連結':   f"https://24h.pchome.com.tw/prod/{p.get('Id', '')}",
            })
        time.sleep(1)
    logging.info(f"PChome 台灣：{len(products)} 筆")
    return products

# ── 美國市場 ───────────────────────────────────────────────────────────────────

def fetch_lg_us():
    """LG 美國官網除濕機"""
    products = []
    r = safe_get("https://www.lg.com/us/dehumidifiers")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('li.item, .product-card, [data-product-id]'):
        name_el  = item.select_one('.product-name, .model-name, h3, h4')
        price_el = item.select_one('.price, .product-price, [class*="price"]')
        model_el = item.select_one('.model-id, [class*="model"]')
        link_el  = item.select_one('a[href]')
        spec_el  = item.select_one('.spec, .capacity, [class*="spec"]')
        name  = clean(name_el.text)  if name_el  else ''
        price = clean(price_el.text) if price_el else ''
        model = clean(model_el.text) if model_el else ''
        spec  = clean(spec_el.text)  if spec_el  else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.lg.com' + href
        if name:
            products.append({
                '生產地': '韓國/全球',
                '品牌':   'LG',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   model,
                '品名':   name,
                '主要規格': spec,
                '單價':   f"USD {price}" if price else '',
                '來源':   'LG 美國官網',
                '連結':   href,
            })
    logging.info(f"LG 美國：{len(products)} 筆")
    return products


def fetch_frigidaire_us():
    """Frigidaire 美國官網除濕機"""
    products = []
    r = safe_get("https://www.frigidaire.com/Home-Comfort/Dehumidifiers/")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-tile, .plp-product, article[class*="product"]'):
        name_el  = item.select_one('h3, h4, .product-name, [class*="title"]')
        price_el = item.select_one('.price, [class*="price"]')
        model_el = item.select_one('.model-number, [class*="model"]')
        link_el  = item.select_one('a[href]')
        name  = clean(name_el.text)  if name_el  else ''
        price = clean(price_el.text) if price_el else ''
        model = clean(model_el.text) if model_el else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.frigidaire.com' + href
        if name:
            products.append({
                '生產地': '美國/中國',
                '品牌':   'Frigidaire',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   model,
                '品名':   name,
                '主要規格': '',
                '單價':   f"USD {price}" if price else '',
                '來源':   'Frigidaire 美國官網',
                '連結':   href,
            })
    logging.info(f"Frigidaire 美國：{len(products)} 筆")
    return products


def fetch_walmart_us():
    """Walmart 美國（公開搜尋 API）"""
    products = []
    r = safe_get(
        "https://www.walmart.com/search",
        params={'q': 'dehumidifier', 'sort': 'new'},
        extra_headers={'Accept': 'text/html'},
    )
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('[data-item-id], [data-automation-id="product"]'):
        name_el  = item.select_one('[itemprop="name"], .product-title, span[class*="title"]')
        price_el = item.select_one('[itemprop="price"], [class*="price"]')
        link_el  = item.select_one('a[href]')
        name  = clean(name_el.text)  if name_el  else ''
        price = clean(price_el.get('content', price_el.text)) if price_el else ''
        href  = link_el['href'] if link_el else ''
        if href and not href.startswith('http'):
            href = 'https://www.walmart.com' + href
        if name:
            products.append({
                '生產地': '美國市場',
                '品牌':   '',
                '上市時間': datetime.now().strftime('%Y-%m'),
                '型號':   '',
                '品名':   name,
                '主要規格': '',
                '單價':   f"USD {price}" if price else '',
                '來源':   'Walmart 美國',
                '連結':   href,
            })
    logging.info(f"Walmart 美國：{len(products)} 筆")
    return products

# ── 歐洲市場 ───────────────────────────────────────────────────────────────────

def fetch_panasonic_eu():
    """Panasonic 歐洲官網除濕機"""
    products = []
    r = safe_get("https://www.panasonic.com/uk/consumer/air-treatment/dehumidifiers.html")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-list__item, .product-card, [class*="product-item"]'):
        name_el  = item.select_one('h3, h4, .product-name, [class*="name"]')
        price_el = item.select_one('.price, [class*="price"]')
        model_el = item.select_one('.model, [class*="model"]')
        link_el  = item.select_one('a[href]')
        spec_el  = item.select_one('.spec, [class*="spec"], .capacity')
        name  = clean(name_el.text)  if name_el  else ''
        price = clean(price_el.text) if price_el else ''
        model = clean(model_el.text) if model_el else ''
        spec  = clean(spec_el.text)  if spec_el  else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.panasonic.com' + href
        if name:
            products.append({
                '生產地': '日本/歐洲',
                '品牌':   'Panasonic',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   model,
                '品名':   name,
                '主要規格': spec,
                '單價':   f"GBP {price}" if price else '',
                '來源':   'Panasonic 歐洲官網',
                '連結':   href,
            })
    logging.info(f"Panasonic 歐洲：{len(products)} 筆")
    return products


def fetch_meaco_uk():
    """Meaco 英國品牌官網（歐洲市場代表性品牌）"""
    products = []
    r = safe_get("https://www.meaco.com/collections/dehumidifiers")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-item, .grid__item, [class*="product"]'):
        name_el  = item.select_one('.product-item__title, h3, h4, [class*="title"]')
        price_el = item.select_one('.price, [class*="price"]')
        link_el  = item.select_one('a[href]')
        name  = clean(name_el.text)  if name_el  else ''
        price = clean(price_el.text) if price_el else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.meaco.com' + href
        if name:
            products.append({
                '生產地': '英國/歐洲',
                '品牌':   'Meaco',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   '',
                '品名':   name,
                '主要規格': '',
                '單價':   f"GBP {price}" if price else '',
                '來源':   'Meaco 英國官網',
                '連結':   href,
            })
    logging.info(f"Meaco 英國：{len(products)} 筆")
    return products

# ── 中國/亞洲製造 ──────────────────────────────────────────────────────────────

def fetch_midea_global():
    """Midea 美的全球官網（中國最大品牌）"""
    products = []
    r = safe_get("https://www.midea.com/global/products/Air_Treatment/Dehumidifier/")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-item, .product-card, [class*="product"]'):
        name_el  = item.select_one('h3, h4, .title, [class*="name"]')
        model_el = item.select_one('.model, [class*="model"]')
        spec_el  = item.select_one('.spec, [class*="capacity"], [class*="spec"]')
        link_el  = item.select_one('a[href]')
        name  = clean(name_el.text)  if name_el  else ''
        model = clean(model_el.text) if model_el else ''
        spec  = clean(spec_el.text)  if spec_el  else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.midea.com' + href
        if name:
            products.append({
                '生產地': '中國大陸',
                '品牌':   'Midea 美的',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   model,
                '品名':   name,
                '主要規格': spec,
                '單價':   '',
                '來源':   'Midea 美的全球官網',
                '連結':   href,
            })
    logging.info(f"Midea 全球：{len(products)} 筆")
    return products


def fetch_made_in_china():
    """Made-in-China.com（中國製造商 B2B 採購平台）"""
    products = []
    r = safe_get(
        "https://www.made-in-china.com/products-search/hot-china-products/Dehumidifier.html",
        extra_headers={'Accept-Language': 'en-US,en;q=0.9'},
    )
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-item, .pro-item, [class*="product"]'):
        name_el    = item.select_one('h4, h3, .pro-name, [class*="name"]')
        price_el   = item.select_one('.price, [class*="price"]')
        company_el = item.select_one('.company-name, [class*="company"]')
        link_el    = item.select_one('a[href]')
        name    = clean(name_el.text)    if name_el    else ''
        price   = clean(price_el.text)   if price_el   else ''
        company = clean(company_el.text) if company_el else ''
        href    = link_el['href']        if link_el    else ''
        if href and not href.startswith('http'):
            href = 'https://www.made-in-china.com' + href
        if name:
            products.append({
                '生產地': '中國大陸',
                '品牌':   company,
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   '',
                '品名':   name,
                '主要規格': '',
                '單價':   f"USD {price}" if price else '',
                '來源':   'Made-in-China',
                '連結':   href,
            })
    logging.info(f"Made-in-China：{len(products)} 筆")
    return products


def fetch_haier_global():
    """Haier 海爾全球官網"""
    products = []
    r = safe_get("https://www.haier.com/global/air-conditioners/dehumidifier/")
    if not r:
        return products
    soup = BeautifulSoup(r.text, 'lxml')
    for item in soup.select('.product-item, [class*="product"]'):
        name_el  = item.select_one('h3, h4, [class*="name"], [class*="title"]')
        model_el = item.select_one('[class*="model"]')
        spec_el  = item.select_one('[class*="spec"], [class*="capacity"]')
        link_el  = item.select_one('a[href]')
        name  = clean(name_el.text)  if name_el  else ''
        model = clean(model_el.text) if model_el else ''
        spec  = clean(spec_el.text)  if spec_el  else ''
        href  = link_el['href']      if link_el  else ''
        if href and not href.startswith('http'):
            href = 'https://www.haier.com' + href
        if name:
            products.append({
                '生產地': '中國大陸',
                '品牌':   'Haier 海爾',
                '上市時間': datetime.now().strftime('%Y'),
                '型號':   model,
                '品名':   name,
                '主要規格': spec,
                '單價':   '',
                '來源':   'Haier 海爾全球官網',
                '連結':   href,
            })
    logging.info(f"Haier 全球：{len(products)} 筆")
    return products

# ── Excel 輸出 ─────────────────────────────────────────────────────────────────
def save_to_excel(all_products: list) -> Path | None:
    if not all_products:
        logging.warning("沒有收集到任何資料")
        return None

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today    = datetime.now().strftime('%Y%m%d')
    filepath = OUTPUT_DIR / f"除濕機資料_{today}.xlsx"
    if filepath.exists():
        ts = datetime.now().strftime('%H%M%S')
        filepath = OUTPUT_DIR / f"除濕機資料_{today}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = f"除濕機_{today}"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(name='微軟正黑體', bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    even_fill = PatternFill("solid", fgColor="EBF3FB")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    thin      = Side(style='thin')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = hdr_align; c.border = border

    for ri, prod in enumerate(all_products, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill
        ws.row_dimensions[ri].height = 20
        for ci, col in enumerate(COLUMNS, 1):
            c = ws.cell(row=ri, column=ci, value=prod.get(col, ''))
            c.fill = fill
            c.font = Font(name='微軟正黑體', size=10)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = border

    for ci, w in enumerate([12, 15, 12, 18, 40, 35, 14, 14, 45], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet("統計摘要")
    ws2['A1'] = f"更新時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2['A2'] = f"總筆數：{len(all_products)} 筆"
    counts: dict = {}
    for p in all_products:
        s = p.get('來源', '未知')
        counts[s] = counts.get(s, 0) + 1
    for i, (src, cnt) in enumerate(counts.items(), 3):
        ws2[f'A{i}'] = f"  {src}：{cnt} 筆"

    wb.save(str(filepath))
    logging.info(f"Excel 已儲存：{filepath}（{len(all_products)} 筆）")
    return filepath

# ── Email 寄送 ─────────────────────────────────────────────────────────────────
def send_email(filepath: Path, total: int, counts: dict):
    if not SMTP_USER or not SMTP_PASS:
        print("  [略過] 未設定 GMAIL_USER / GMAIL_APP_PASS，跳過寄信")
        logging.info("Email 未設定，略過寄信")
        return

    today   = datetime.now().strftime('%Y-%m-%d')
    subject = f"除濕機全球最新資料 {today}（共 {total} 筆）"
    body_lines = [
        f"您好，",
        f"",
        f"以下是 {today} 全球除濕機最新資料，共收集 {total} 筆，請見附件 Excel。",
        f"",
        f"各來源筆數：",
    ] + [f"  • {src}：{cnt} 筆" for src, cnt in counts.items()] + [
        f"",
        f"此郵件由 GitHub Actions 自動寄出，每天 08:00（台灣時間）執行。",
    ]

    msg = MIMEMultipart()
    msg['From']    = SMTP_USER
    msg['To']      = RECIPIENT
    msg['Subject'] = subject
    msg.attach(MIMEText("\n".join(body_lines), 'plain', 'utf-8'))

    with open(filepath, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filepath.name}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, RECIPIENT, msg.as_string())
        print(f"  Email 已寄送至 {RECIPIENT}")
        logging.info(f"Email 寄送成功 → {RECIPIENT}")
    except Exception as e:
        print(f"  [警告] Email 寄送失敗：{e}")
        logging.error(f"Email 寄送失敗：{e}")

# ── 主程式 ────────────────────────────────────────────────────────────────────
def main():
    setup_logging()
    logging.info("=== 除濕機資料收集開始 ===")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 開始收集全球除濕機資料...")

    sources = [
        # 台灣
        ("PChome 台灣",       fetch_pchome_taiwan),
        # 美國
        ("LG 美國官網",       fetch_lg_us),
        ("Frigidaire 美國",   fetch_frigidaire_us),
        ("Walmart 美國",      fetch_walmart_us),
        # 歐洲
        ("Panasonic 歐洲",    fetch_panasonic_eu),
        ("Meaco 英國",        fetch_meaco_uk),
        # 中國/亞洲
        ("Midea 美的全球",    fetch_midea_global),
        ("Haier 海爾全球",    fetch_haier_global),
        ("Made-in-China",     fetch_made_in_china),
    ]

    all_products = []
    for name, func in sources:
        print(f"  正在抓取 {name} ...", end='', flush=True)
        prods = func()
        all_products.extend(prods)
        print(f" {len(prods)} 筆")
        time.sleep(2)

    print(f"\n共收集 {len(all_products)} 筆，正在儲存 Excel ...")
    filepath = save_to_excel(all_products)

    if filepath:
        print(f"檔案：{filepath}")
        counts = {}
        for p in all_products:
            s = p.get('來源', '未知')
            counts[s] = counts.get(s, 0) + 1
        print("正在寄送 Email ...")
        send_email(filepath, len(all_products), counts)
    else:
        print("警告：未能收集到任何資料")

    logging.info("=== 除濕機資料收集結束 ===")


if __name__ == "__main__":
    main()
